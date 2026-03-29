"""
Siemens TIA Portal DB Instrument Configuration Checker
=======================================================
Reads testBrew.xlsx + SysConfig_Lib_Brew.db,
compares ENABLE/EXIST/VLV_W_FB settings for each configured family,
produces:
  - SysConfig_Lib_Brew_updated.db  (corrected DB for TIA Portal import)
  - InstrumentStatusComparison.xlsx (full comparison report)

Usage:
  python main.py

Sustainable: update the Excel family column or swap the .db file and re-run.
"""

import re
import sys
import openpyxl
import xlsxwriter
from collections import defaultdict

# Force UTF-8 output on Windows
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Configuration ──────────────────────────────────────────────────────────────
DB_PATH    = "SysConfig_Lib_Brew.db"
EXCEL_PATH = "testBrew.xlsx"
OUT_DB     = "SysConfig_Lib_Brew_updated.db"
OUT_XLSX   = "InstrumentStatusComparison.xlsx"

# Map Excel column header → (MachineConfig index, 1-based column index in Excel)
FAMILIES = {
    "BREW 350": {"mc_idx": 7, "col_idx": 4},   # column D
    "BREW 450": {"mc_idx": 8, "col_idx": 5},   # column E
}

# ── Tag path map: Excel label → (unit, em, db_field, is_vlv) ─────────────────
# is_vlv=True means STAT VLV type (has VLV_W_FB bit)
TAG_MAP = {
    # EM-100 (Unit Sep) — Utilities
    "L-GN":        ("Unit Sep",     "EM - 100",  "L-GREEN",       False),
    "L-YL":        ("Unit Sep",     "EM - 100",  "L-YELLOW",      False),
    "L-RD":        ("Unit Sep",     "EM - 100",  "L-RED",         False),
    "L-BUZZ":      ("Unit Sep",     "EM - 100",  "L-BUZZ",        False),
    "EMS":         ("Unit Sep",     "EM - 100",  "EMS",           False),
    "UPS":         ("Unit Sep",     "EM - 100",  "UPS",           False),
    "eWON":        ("Unit Sep",     "EM - 100",  "EWON",          False),
    "PS502-1":     ("Unit Sep",     "EM - 100",  "PS502-1",       False),
    "LT-BT":       ("Unit Sep",     "EM - 100",  "LT BUF",        False),
    # EM-400 (Unit Sep) — Cooling liquid
    "AV409-1":     ("Unit Sep",     "EM - 400",  "AV409-1",       True),
    "FCV409-1":    ("Unit Sep",     "EM - 400",  "FCV409-1",      False),
    "FIS409-1":    ("Unit Sep",     "EM - 400",  "FIS409-1",      False),
    "FIT409-1":    ("Unit Sep",     "EM - 400",  "FIT409-1",      False),
    "TT41X-1":     ("Unit Sep",     "EM - 400",  "TT41X-1",       False),
    # EM-460 (Unit Sep) — Liquid seal
    "AV460d-1":    ("Unit Sep",     "EM - 460",  "AV460d-1",      True),
    # EM-600 (Unit Sep) — Sealing liquids
    "AV615-1":     ("Unit Sep",     "EM - 600",  "AV615-1",       True),
    "AV63X/441-1": ("Unit Sep",     "EM - 600",  "AV63X/441-1",   True),
    "FIS615-1":    ("Unit Sep",     "EM - 600",  "FIS615-1",      False),
    "FIS630-1":    ("Unit Sep",     "EM - 600",  "FIS630-1",      False),
    "FIS635-1":    ("Unit Sep",     "EM - 600",  "FIS635-1",      False),
    # EM-701 (Unit Sep) — Main drive
    "M701":        ("Unit Sep",     "EM - 701",  "M701",          False),
    # EM-730 (Unit Sep) — Temperature supervision
    "TS730":       ("Unit Sep",     "EM - 730",  "TS730",         False),
    # EM-740 (Unit Sep) — Speed supervision
    "P+F PS":      ("Unit Sep",     "EM - 740",  "P+F PS",        False),
    "P+F OS":      ("Unit Sep",     "EM - 740",  "P+F OS",        False),
    "P+F ERROR":   ("Unit Sep",     "EM - 740",  "P+F ERROR",     False),
    # EM-750 (Unit Sep) — Vibration supervision
    "YT75X":       ("Unit Sep",     "EM - 750",  "YT75X",         False),
    # EM-201 (Unit Process) — Product inlet
    "AV201-1":     ("Unit Process", "EM - 201",  "AV201-1",       True),
    "FCV201-1":    ("Unit Process", "EM - 201",  "FCV201-1",      False),
    "FIT201-1":    ("Unit Process", "EM - 201",  "FIT201-1",      False),
    "LA201-1":     ("Unit Process", "EM - 201",  "LA201-1",       False),
    "PIT201-1":    ("Unit Process", "EM - 201",  "PIT201-1",      False),
    "QT201-1":     ("Unit Process", "EM - 201",  "QT201-1",       False),
    "QT201-1_OK":  ("Unit Process", "EM - 201",  "QT201-1_OK",    False),
    # EM-212 (Unit Process) — Bypass
    "AV212-1":     ("Unit Process", "EM - 212",  "AV212-1",       True),
    # EM-22X (Unit Process) — LP outlet
    "AV220-1":     ("Unit Process", "EM - 22X",  "AV220-1",       True),
    "AV460-1":     ("Unit Process", "EM - 22X",  "AV460-1",       True),
    "FIT22X-1":    ("Unit Process", "EM - 22X",  "FIT22X-1",      False),
    "PIT22X-1":    ("Unit Process", "EM - 22X",  "PIT22X-1",      False),
    "QT220-1":     ("Unit Process", "EM - 22X",  "QT220-1",       False),
    "PCV22X-1":    ("Unit Process", "EM - 22X",  "PCV22X-1",      False),
    "QT220-1_OK":  ("Unit Process", "EM - 22X",  "QT220-1_OK",    False),
    "LA220-1":     ("Unit Process", "EM - 22X",  "LA220-1",       False),
    # EM-22Y (Unit Process) — HP outlet
    "AV221-1":     ("Unit Process", "EM - 22Y",  "AV221-1",       True),
    "AV461-1":     ("Unit Process", "EM - 22Y",  "AV461-1",       True),
    "LA221-1":     ("Unit Process", "EM - 22Y",  "LA221-1",       False),
    "PCV22Y-1":    ("Unit Process", "EM - 22Y",  "PCV22Y-1",      False),
    "PIT22Y-1":    ("Unit Process", "EM - 22Y",  "PIT22Y-1",      False),
    # EM-340 (Unit Process) — Safety water
    "AV340-1":     ("Unit Process", "EM - 340",  "AV340-1",       True),
    "AV340-2":     ("Unit Process", "EM - 340",  "AV340-2",       True),
    "AV460-2":     ("Unit Process", "EM - 340",  "AV460-2",       True),
    "PIS340-1":    ("Unit Process", "EM - 340",  "PIS340-1",      False),
    # EM-222 (Unit Dch) — SRU / solids outlet
    "AV222-1":     ("Unit Dch",     "EM - 222",  "AV222-1",       True),
    "LS222-1":     ("Unit Dch",     "EM - 222",  "LS222-1",       False),
    "LS222-2":     ("Unit Dch",     "EM - 222",  "LS222-2",       False),
    "LS222-3":     ("Unit Dch",     "EM - 222",  "LS222-3",       False),
    "P222-1":      ("Unit Dch",     "EM - 222",  "P222-1",        False),
    "TS222-1":     ("Unit Dch",     "EM - 222",  "TS222-1",       False),
    "TT222-1":     ("Unit Dch",     "EM - 222",  "TT222-1",       False),
    # EM-300 (Unit Dch) — Flushing
    "AV301-1":     ("Unit Dch",     "EM - 300",  "AV301-1",       True),
    "AV302-1":     ("Unit Dch",     "EM - 300",  "AV302-1",       True),
    "AV303-1":     ("Unit Dch",     "EM - 300",  "AV303-1",       True),
    "AV304-1":     ("Unit Dch",     "EM - 300",  "AV304-1",       True),
    "AV305-1":     ("Unit Dch",     "EM - 300",  "AV305-1",       True),
    "AV306-1":     ("Unit Dch",     "EM - 300",  "AV306-1",       True),
    "AV320-1":     ("Unit Dch",     "EM - 300",  "AV320-1",       True),
    "AV635-1":     ("Unit Dch",     "EM - 300",  "AV635-1",       True),
    # EM-375 (Unit Dch) — Op. water
    "AV37X-1":     ("Unit Dch",     "EM - 375",  "AV37X-1",       True),
    "PIS375-1":    ("Unit Dch",     "EM - 375",  "PIS375-1",      False),
    # EM-506 (Unit Dch) — Discharge system
    "AV543-1":     ("Unit Dch",     "EM - 506",  "AV543-1",       True),
    "IP506A":      ("Unit Dch",     "EM - 506",  "IP506a",        False),
    "AVDCH":       ("Unit Dch",     "EM - 506",  "AVDCH",         True),
    "PV376-1":     ("Unit Dch",     "EM - 506",  "PV376-1",       True),
    "PV376-2":     ("Unit Dch",     "EM - 506",  "PV376-2",       True),
    # Options — EM-P201 (Feed pump)
    "P201-1":      ("Options",      "EM - P201", "P201-1",        False),
    # Options — EM-208 (Blending)
    "AV208-1":     ("Options",      "EM - 208",  "AV208-1",       True),
    "FCV208-1":    ("Options",      "EM - 208",  "FCV208-1",      False),
    "FIT208-1":    ("Options",      "EM - 208",  "FIT208-1",      False),
    "P208-1":      ("Options",      "EM - 208",  "P208-1",        False),
    "QT220-2":     ("Options",      "EM - 208",  "QT220-2",       False),
    "QT220-2_OK":  ("Options",      "EM - 208",  "QT220-2_OK",    False),
    # Options — EM-210 (Recirculation)
    "AV210-1":     ("Options",      "EM - 210",  "AV210-1",       True),
    "FCV210-1":    ("Options",      "EM - 210",  "FCV210-1",      False),
    # Options — EM-430 (Cooling recirculation)
    "AV430-1":     ("Options",      "EM - 430",  "AV430-1",       True),
    "FCV430-1":    ("Options",      "EM - 430",  "FCV430-1",      False),
    "P406-1":      ("Options",      "EM - 430",  "P406-1",        False),
    "TT406-1":     ("Options",      "EM - 430",  "TT406-1",       False),
    # Options — EM-463 (eMotion)
    "AV305-2":     ("Options",      "EM - 463",  "AV305-2",       True),
    "AV463-1":     ("Options",      "EM - 463",  "AV463-1",       True),
    "AV463-2":     ("Options",      "EM - 463",  "AV463-2",       True),
    "AV470-1":     ("Options",      "EM - 463",  "AV470-1",       True),
    "AV660-1":     ("Options",      "EM - 463",  "AV660-1",       True),
    "AV660-2":     ("Options",      "EM - 463",  "AV660-2",       True),
    "LT463-1":     ("Options",      "EM - 463",  "LT463-1",       False),
    "PIT463-1":    ("Options",      "EM - 463",  "PIT463-1",      False),
    "P463-1":      ("Options",      "EM - 463",  "P463-1",        False),
}

# ── Low-level helpers ─────────────────────────────────────────────────────────

def norm_bool(val_str):
    return val_str.strip().lower() in ("true",)


def path_key(unit, em, field):
    return f"{unit}|{em}|{field}"


def extract_tuple_content(text, start_pos):
    """Return content between matching parens starting at start_pos (must be '(')."""
    assert text[start_pos] == "(", f"Expected '(' at {start_pos}"
    depth = 0
    i = start_pos
    while i < len(text):
        c = text[i]
        if c == "(":
            depth += 1
        elif c == ")":
            depth -= 1
            if depth == 0:
                return text[start_pos + 1 : i]
        i += 1
    return text[start_pos + 1:]


def parse_cw_bits(cw_content):
    """
    Parse the 16-bit CW tuple content string like:
      '(), (), (), True, (), (), TRUE, TRUE, (), ..., True, (), ()'
    Returns list of 16 booleans.
    """
    bits = []
    i = 0
    s = cw_content
    while i < len(s):
        c = s[i]
        if c in " \t\r\n,":
            i += 1
        elif c == "(":
            # () element → False
            j = s.find(")", i)
            bits.append(False)
            i = j + 1
        elif c in "TtFf":
            # Bool token
            j = i
            while j < len(s) and s[j].isalpha():
                j += 1
            word = s[i:j].upper()
            bits.append(word == "TRUE")
            i = j
        else:
            # Number or unknown — skip to next comma
            j = i
            while j < len(s) and s[j] not in ",)":
                j += 1
            bits.append(False)
            i = j
    return bits


def make_db_line(mc_idx, unit, em, field, prop, value):
    """Generate a MachineConfig DB assignment line."""
    bool_str = "True" if value else "False"
    def q(s):
        if re.search(r"[\s\-\+/]", s):
            return f'"{s}"'
        return s
    field_q = q(field)
    return f'   MachineConfig[{mc_idx}]."{unit}"."{em}".{field_q}.CW.{prop} := {bool_str};\r\n'


# ── Step 1: Parse Excel ───────────────────────────────────────────────────────

def parse_excel(path):
    """
    Returns: {family_name: {excel_tag: {enable, exist, vlv_w_fb, function, is_vlv, ...}}}
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Find header row, family columns, and optional "Option" descriptor column
    family_cols = {}   # {col_idx (1-based): family_name}
    option_col  = None # col_idx (1-based) for the "Option" descriptor column
    header_row  = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                if v in FAMILIES:
                    family_cols[cell.column] = v
                    header_row = cell.row
                elif v.lower() == "option":
                    option_col = cell.column

    if not family_cols:
        raise ValueError("No family columns found in Excel")

    # Read all data rows
    raw_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        tag_cell  = row[1]   # column B
        func_cell = row[2]   # column C
        tag_label = tag_cell.value
        func      = func_cell.value if func_cell.value else ""

        if not tag_label:
            continue
        tag_label = str(tag_label).strip()
        func      = str(func).strip()

        if tag_label.upper() in ("SPARE", "DUMMY", "") or func.upper() in ("SPARE", "DUMMY"):
            continue

        fam_vals = {}
        for col_idx, fam_name in family_cols.items():
            cell    = row[col_idx - 1]
            raw_val = cell.value
            comment = None
            if cell.comment:
                # Extract plain last-line of comment (skip boilerplate)
                lines = [ln.strip() for ln in cell.comment.text.splitlines() if ln.strip()]
                # Last meaningful line
                for ln in reversed(lines):
                    if not any(x in ln.lower() for x in ("http", "version", "learn", "excel", "comment:", "threaded")):
                        comment = ln
                        break
            fam_vals[fam_name] = {"val": raw_val, "comment": comment}

        # Read option descriptor column if present (e.g. "Option" column = BLENDING, SRU...)
        opt_desc = ""
        if option_col:
            opt_desc = str(row[option_col - 1].value or "").strip().upper()

        raw_rows.append((tag_label, func, fam_vals, opt_desc))

    # Identify option-type tags. Two conditions (either is sufficient):
    #  1. Excel "Option" column (or family cell) contains an option descriptor
    #  2. The tag's DB path is under unit "Options" (any EM found under Options in the DB)
    # For option-type tags, 'o' → ENABLE=False, EXIST=False, VLV_W_FB=False.
    OPTION_MARKERS = {"BLENDING", "RECIRCULATION", "COOLING RECIRCULATION", "SRU", "FEED PUMP"}
    option_tags = set()
    # Condition 1 — Excel Option column or legacy family-cell descriptor
    for _tag, _func, _fam_vals, _opt_desc in raw_rows:
        if _opt_desc in OPTION_MARKERS:
            option_tags.add(_tag)
        else:
            for _fv in _fam_vals.values():
                if str(_fv.get("val") or "").strip().upper() in OPTION_MARKERS:
                    option_tags.add(_tag)
                    break
    # Condition 2 — tag lives under the "Options" unit in the DB (TAG_MAP lookup)
    for _tag_key, _tag_info in TAG_MAP.items():
        if _tag_info[0] == "Options":
            option_tags.add(_tag_key)

    # Aggregate rows by tag into valve (multi-row) vs non-valve (single row)
    valve_rows = defaultdict(lambda: {"fb_opn": {}, "fb_cls": {}, "act": {}})
    non_valve_rows = {}   # tag → {fam_vals, func}

    AV_FUNCS = {"FB OPN", "FB CLS", "ACT"}

    for tag_label, func, fam_vals, _opt in raw_rows:
        func_up = func.upper().strip()
        is_av = tag_label.upper().startswith("AV")

        if is_av and func_up in AV_FUNCS:
            if func_up == "FB OPN":
                valve_rows[tag_label]["fb_opn"] = fam_vals
            elif func_up == "FB CLS":
                valve_rows[tag_label]["fb_cls"] = fam_vals
            elif func_up == "ACT":
                valve_rows[tag_label]["act"] = fam_vals
        else:
            # Non-valve (or AV with unexpected function): direct row
            if tag_label not in non_valve_rows:
                non_valve_rows[tag_label] = {"fam_vals": {}, "func": func}
            for fam, fv in fam_vals.items():
                if fam not in non_valve_rows[tag_label]["fam_vals"]:
                    non_valve_rows[tag_label]["fam_vals"][fam] = fv
                elif non_valve_rows[tag_label]["fam_vals"][fam]["val"] is None and fv["val"] is not None:
                    non_valve_rows[tag_label]["fam_vals"][fam] = fv

    def xlval_to_enable_exist(raw_val, is_option=False):
        if raw_val is None:
            return False, False
        v = str(raw_val).strip().lower()
        if v == "x":
            return True, True
        elif v == "o":
            if is_option:
                return False, False   # option module not selected → both off
            return True, False
        return False, False

    def has_fb(fam_vals, fam_name, is_option=False):
        fv = fam_vals.get(fam_name, {}).get("val")
        if fv is None:
            return False
        v = str(fv).strip().lower()
        if is_option and v == "o":
            return False   # option module 'o' = not installed → no feedback
        return v in ("x", "o")

    families_data = {fam: {} for fam in family_cols.values()}

    # Process valves
    for tag_label, rows in valve_rows.items():
        for fam_name in family_cols.values():
            act_fv   = rows["act"].get(fam_name, {"val": None, "comment": None})
            fb_opn   = rows["fb_opn"].get(fam_name, {"val": None})
            fb_cls   = rows["fb_cls"].get(fam_name, {"val": None})

            is_opt   = tag_label in option_tags
            enable, exist = xlval_to_enable_exist(act_fv.get("val"), is_option=is_opt)
            vlv_w_fb = has_fb(rows["fb_opn"], fam_name, is_option=is_opt) or has_fb(rows["fb_cls"], fam_name, is_option=is_opt)

            resolved = act_fv.get("comment") or tag_label

            families_data[fam_name][tag_label] = {
                "enable":   enable,
                "exist":    exist,
                "vlv_w_fb": vlv_w_fb,
                "resolved_tag": resolved,
                "label":    tag_label,
                "function": "AV Valve",
                "is_vlv":   True,
                "raw_act":  act_fv.get("val"),
                "raw_fb":   fb_opn.get("val") or fb_cls.get("val"),
            }

    # Process non-valves
    for tag_label, info in non_valve_rows.items():
        for fam_name in family_cols.values():
            fv = info["fam_vals"].get(fam_name, {"val": None, "comment": None})
            raw_val = fv.get("val")
            comment = fv.get("comment")
            enable, exist = xlval_to_enable_exist(raw_val, is_option=(tag_label in option_tags))
            resolved = comment if comment else tag_label

            families_data[fam_name][tag_label] = {
                "enable":   enable,
                "exist":    exist,
                "vlv_w_fb": False,
                "resolved_tag": resolved,
                "label":    tag_label,
                "function": info["func"],
                "is_vlv":   False,
                "raw_val":  raw_val,
            }

    return families_data


# ── Step 2: Parse DB type defaults ───────────────────────────────────────────

def parse_type_defaults(db_text):
    """
    Parse TYPE ... END_TYPE blocks and extract default ENABLE/EXIST/VLV_W_FB
    for every instrument (STAT VLV/DI/DO/AI/AO/MTR/VFD) in each EM type.
    Returns: {path_key: {default_enable, default_exist, default_vlv_w_fb, ...}}
    """
    EM_UNIT_MAP = {
        "EM - 100": "Unit Sep",
        "EM - 400": "Unit Sep",
        "EM - 420": "Unit Sep",
        "EM - 460": "Unit Sep",
        "EM - 600": "Unit Sep",
        "EM - 701": "Unit Sep",
        "EM - 730": "Unit Sep",
        "EM - 740": "Unit Sep",
        "EM - 750": "Unit Sep",
        "EM - 508": "Unit Sep",
        "EM - 201": "Unit Process",
        "EM - 206": "Unit Process",
        "EM - 212": "Unit Process",
        "EM - 22X": "Unit Process",
        "EM - 22Y": "Unit Process",
        "EM - 340": "Unit Process",
        "EM - 222": "Unit Dch",
        "EM - 300": "Unit Dch",
        "EM - 375": "Unit Dch",
        "EM - 506": "Unit Dch",
        "EM - P201": "Options",
        "EM - 208": "Options",
        "EM - 210": "Options",
        "EM - 430": "Options",
        "EM - 463": "Options",
        "EM - 551": "Options",
        "EM - PRU1": "Options",
        "EM - PRU2": "Options",
    }
    STAT_TYPES = {"STAT VLV", "STAT DI", "STAT DO", "STAT AI", "STAT AO", "STAT MTR", "STAT VFD"}

    defaults = {}

    type_pat = re.compile(r'TYPE\s+"([^"]+)".*?END_TYPE', re.DOTALL)

    # Pattern to find field start: name : "STAT XXX" :=
    field_hdr_pat = re.compile(
        r'"?([A-Za-z0-9_\-\+\./]+(?:\([a-zA-Z0-9]+\))?)"?'  # field name
        r'(?:\s*\{[^}]*\})?\s*:\s*"(STAT [A-Z]+)"\s*:=\s*',  # : "STAT XX" :=
        re.DOTALL
    )

    for type_match in type_pat.finditer(db_text):
        type_name = type_match.group(1)
        if type_name not in EM_UNIT_MAP:
            continue

        unit     = EM_UNIT_MAP[type_name]
        em       = type_name
        body     = type_match.group(0)
        body_start = type_match.start()  # absolute position in db_text

        for fm in field_hdr_pat.finditer(body):
            field_raw = fm.group(1).strip()
            stat_type = fm.group(2).strip()
            if stat_type not in STAT_TYPES:
                continue

            is_vlv = (stat_type == "STAT VLV")

            # After ':=' find the outer '(' of the default tuple
            after_eq = fm.end()
            # Skip whitespace
            p = after_eq
            while p < len(body) and body[p] in " \t\r\n":
                p += 1
            if p >= len(body) or body[p] != "(":
                continue

            # Extract outer tuple: ('tag', (cw_bits), ...)
            outer_content = extract_tuple_content(body, p)

            # Find the tag string (first quoted arg)
            q1 = outer_content.find("'")
            if q1 == -1:
                continue
            q2 = outer_content.find("'", q1 + 1)
            if q2 == -1:
                continue

            # After 'tag', the next '(' is the CW tuple
            after_tag = q2 + 1
            cw_paren = outer_content.find("(", after_tag)
            if cw_paren == -1:
                continue

            cw_content = extract_tuple_content(outer_content, cw_paren)
            bits = parse_cw_bits(cw_content)

            enable_def  = bits[3]  if len(bits) > 3  else False
            exist_def   = bits[13] if len(bits) > 13 else False
            vlv_w_fb_def = bits[15] if (is_vlv and len(bits) > 15) else False

            pk = path_key(unit, em, field_raw)
            defaults[pk] = {
                "default_enable":    enable_def,
                "default_exist":     exist_def,
                "default_vlv_w_fb":  vlv_w_fb_def,
                "unit": unit, "em": em, "field": field_raw,
                "is_vlv": is_vlv,
            }

    return defaults


# ── Step 3: Parse MachineConfig[N] overrides ─────────────────────────────────

def parse_mc_overrides(db_text, mc_idx):
    """
    Returns: {path_key: {enable?, exist?, vlv_w_fb?}} from BEGIN section.
    """
    overrides = defaultdict(dict)

    pat = re.compile(
        r'MachineConfig\[' + str(mc_idx) + r'\]\.'
        r'"([^"]+)"\."([^"]+)"\.'
        r'(?:"([^"]+)"|([A-Za-z0-9_\-]+))\.'
        r'CW\.(ENABLE|EXIST|VLV_W_FB)\s*:=\s*(True|true|False|false)\s*;',
        re.IGNORECASE
    )

    for m in pat.finditer(db_text):
        unit  = m.group(1)
        em    = m.group(2)
        field = m.group(3) if m.group(3) else m.group(4)
        prop  = m.group(5).upper()
        val   = norm_bool(m.group(6))
        pk    = path_key(unit, em, field)
        if prop == "ENABLE":
            overrides[pk]["enable"] = val
        elif prop == "EXIST":
            overrides[pk]["exist"] = val
        elif prop == "VLV_W_FB":
            overrides[pk]["vlv_w_fb"] = val

    return dict(overrides)


# ── Step 4: Compute effective values ─────────────────────────────────────────

def compute_effective(type_defaults, mc_overrides, excel_tag):
    """
    Returns (eff_enable, eff_exist, eff_vlv_w_fb, found_in_db).
    """
    if excel_tag not in TAG_MAP:
        return False, False, False, False

    unit, em, field, is_vlv = TAG_MAP[excel_tag]
    pk = path_key(unit, em, field)

    if pk in type_defaults:
        def_en = type_defaults[pk]["default_enable"]
        def_ex = type_defaults[pk]["default_exist"]
        def_vf = type_defaults[pk]["default_vlv_w_fb"]
    else:
        def_en, def_ex, def_vf = False, False, False

    ov     = mc_overrides.get(pk, {})
    eff_en = ov.get("enable", def_en)
    eff_ex = ov.get("exist",  def_ex)
    eff_vf = ov.get("vlv_w_fb", def_vf)

    found = (pk in type_defaults) or (pk in mc_overrides)
    return eff_en, eff_ex, eff_vf, found


# ── Step 5: Build comparison rows ────────────────────────────────────────────

def build_comparison(excel_data, type_defaults, mc_overrides_all):
    """Build list of comparison dicts for all instruments across all families."""
    rows = []

    # Ordered unique tags
    seen = {}
    for fam_name, fam_data in excel_data.items():
        for tag_label, info in fam_data.items():
            if tag_label not in seen:
                seen[tag_label] = info.get("function", "")

    for tag_label, function in seen.items():
        unit = TAG_MAP[tag_label][0] if tag_label in TAG_MAP else "N/A"
        em   = TAG_MAP[tag_label][1] if tag_label in TAG_MAP else "N/A"
        row = {"Tag Label": tag_label, "Function": function, "Unit": unit, "EM": em}

        in_map = tag_label in TAG_MAP
        is_vlv = TAG_MAP[tag_label][3] if in_map else False

        for fam_name, fam_cfg in FAMILIES.items():
            mc_idx      = fam_cfg["mc_idx"]
            fam_data    = excel_data.get(fam_name, {})
            mc_overrides = mc_overrides_all.get(mc_idx, {})
            fam_short   = fam_name.replace(" ", "")

            if tag_label in fam_data:
                inst    = fam_data[tag_label]
                exp_en  = inst["enable"]
                exp_ex  = inst["exist"]
                exp_vf  = inst["vlv_w_fb"] if is_vlv else False
                raw_disp = (
                    str(inst.get("raw_act", "")).upper() or "empty"
                    if inst.get("is_vlv")
                    else str(inst.get("raw_val", "")).upper() or "empty"
                )
            else:
                exp_en, exp_ex, exp_vf = False, False, False
                raw_disp = "N/A"

            if in_map:
                eff_en, eff_ex, eff_vf, found = compute_effective(
                    type_defaults, mc_overrides, tag_label
                )
            else:
                eff_en, eff_ex, eff_vf, found = False, False, False, False

            en_match = (exp_en == eff_en)
            ex_match = (exp_ex == eff_ex)
            vf_match = (exp_vf == eff_vf) if (in_map and is_vlv) else True

            # Determine action
            if not in_map:
                action = "NOT IN TAG_MAP"
            elif not found:
                action = "NOT FOUND IN DB"
            elif not en_match or not ex_match or (is_vlv and not vf_match):
                parts = []
                if not en_match:
                    parts.append(f"ENABLE: DB={eff_en}->{exp_en}")
                if not ex_match:
                    parts.append(f"EXIST: DB={eff_ex}->{exp_ex}")
                if is_vlv and not vf_match:
                    parts.append(f"VLV_W_FB: DB={eff_vf}->{exp_vf}")
                action = "UPDATE: " + ", ".join(parts) if parts else "OK"
            else:
                action = "OK"

            row[f"{fam_short}_Excel"]   = raw_disp
            row[f"{fam_short}_ExpEN"]   = exp_en
            row[f"{fam_short}_ExpEX"]   = exp_ex
            row[f"{fam_short}_ExpVF"]   = exp_vf if is_vlv else "N/A"
            row[f"{fam_short}_DbEN"]    = eff_en
            row[f"{fam_short}_DbEX"]    = eff_ex
            row[f"{fam_short}_DbVF"]    = eff_vf if is_vlv else "N/A"
            row[f"{fam_short}_Match"]   = "OK" if (en_match and ex_match and vf_match) else "MISMATCH"
            row[f"{fam_short}_Action"]  = action

        rows.append(row)

    return rows


# ── Step 6: Generate updated DB ──────────────────────────────────────────────

def generate_updated_db(db_text, excel_data, type_defaults, mc_overrides_all):
    """
    Returns updated DB text.
    For each family:
      - Strips all MachineConfig[N] CW.ENABLE/EXIST/VLV_W_FB override lines
      - Inserts fresh minimal-patch CW lines based on Excel
    """
    lines = db_text.splitlines(keepends=True)

    # Find BEGIN and END_DATA_BLOCK
    begin_idx = end_idx = None
    for i, line in enumerate(lines):
        s = line.strip()
        if s == "BEGIN":
            begin_idx = i
        if s == "END_DATA_BLOCK":
            end_idx = i

    if begin_idx is None or end_idx is None:
        print("WARNING: Could not find BEGIN/END_DATA_BLOCK")
        return db_text

    # Build CW-line-removal patterns for each family
    remove_pats = {}
    for fam_name, fam_cfg in FAMILIES.items():
        mc_idx = fam_cfg["mc_idx"]
        remove_pats[mc_idx] = re.compile(
            r'MachineConfig\[' + str(mc_idx) + r'\]\.'
            r'"[^"]+"\."[^"]+"\.'
            r'(?:"[^"]+"|[A-Za-z0-9_\-]+)\.'
            r'CW\.(ENABLE|EXIST|VLV_W_FB)\s*:=',
            re.IGNORECASE
        )

    # Find the last line for each MC block (to know where to insert new lines)
    mc_last_idx = {}
    for i in range(begin_idx + 1, end_idx):
        for mc_idx in FAMILIES[list(FAMILIES.keys())[0]]["mc_idx"], FAMILIES[list(FAMILIES.keys())[1]]["mc_idx"]:
            if re.match(r'\s*MachineConfig\[' + str(mc_idx) + r'\]', lines[i]):
                mc_last_idx[mc_idx] = i

    # Build new CW lines for each MC
    new_cw = {}
    for fam_name, fam_cfg in FAMILIES.items():
        mc_idx     = fam_cfg["mc_idx"]
        fam_data   = excel_data.get(fam_name, {})
        mc_lines   = []

        for excel_tag, info in fam_data.items():
            if excel_tag not in TAG_MAP:
                continue
            unit, em, field, is_vlv = TAG_MAP[excel_tag]
            pk = path_key(unit, em, field)

            td = type_defaults.get(pk, {})
            def_en = td.get("default_enable", False)
            def_ex = td.get("default_exist",  False)
            def_vf = td.get("default_vlv_w_fb", False)

            exp_en = info["enable"]
            exp_ex = info["exist"]
            exp_vf = info["vlv_w_fb"] if is_vlv else False

            if exp_en != def_en:
                mc_lines.append(make_db_line(mc_idx, unit, em, field, "ENABLE", exp_en))
            if exp_ex != def_ex:
                mc_lines.append(make_db_line(mc_idx, unit, em, field, "EXIST", exp_ex))
            if is_vlv and exp_vf != def_vf:
                mc_lines.append(make_db_line(mc_idx, unit, em, field, "VLV_W_FB", exp_vf))

        new_cw[mc_idx] = sorted(mc_lines)

    # Rebuild lines
    mc_inserted = {mc_idx: False for mc_idx in new_cw}
    output = []

    for i in range(begin_idx + 1):
        output.append(lines[i])

    for i in range(begin_idx + 1, end_idx):
        line = lines[i]
        skip = any(pat.search(line) for pat in remove_pats.values())
        if not skip:
            output.append(line)
        # Insert new CW lines after the last line of each MC block
        for mc_idx, last_i in mc_last_idx.items():
            if i == last_i and not mc_inserted.get(mc_idx, True):
                mc_inserted[mc_idx] = True
                if new_cw.get(mc_idx):
                    output.append(
                        f'   // === CW auto-patched for MachineConfig[{mc_idx}] ===\r\n'
                    )
                    for cw_line in new_cw[mc_idx]:
                        output.append(cw_line)

    for i in range(end_idx, len(lines)):
        output.append(lines[i])

    return "".join(output)


# ── Step 7: Write comparison Excel ───────────────────────────────────────────

def write_comparison_xlsx(comparison_rows, out_path):
    wb = xlsxwriter.Workbook(out_path)
    ws = wb.add_worksheet("Instrument Comparison")

    hdr_fmt   = wb.add_format({"bold": True, "bg_color": "#003366", "font_color": "white",
                                "border": 1, "align": "center", "valign": "vcenter", "text_wrap": True})
    ok_fmt    = wb.add_format({"bg_color": "#C6EFCE", "border": 1, "align": "center"})
    fail_fmt  = wb.add_format({"bg_color": "#FFC7CE", "border": 1, "align": "center"})
    warn_fmt  = wb.add_format({"bg_color": "#FFEB9C", "border": 1, "align": "center"})
    norm_fmt  = wb.add_format({"border": 1})
    center_fmt= wb.add_format({"border": 1, "align": "center"})
    title_fmt = wb.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E79",
                                "font_color": "white", "align": "center", "valign": "vcenter"})
    sub_fmt   = wb.add_format({"bold": True, "bg_color": "#2E75B6", "font_color": "white",
                                "border": 1, "align": "center"})

    fam_names = list(FAMILIES.keys())
    COLS_PER_FAM = 9
    FIXED_COLS = 4  # Tag Label, Function, Unit, EM Module
    total_cols = FIXED_COLS + len(fam_names) * COLS_PER_FAM

    # Title
    ws.merge_range(0, 0, 0, total_cols - 1,
                   "TIA Portal DB Instrument Status Comparison Report", title_fmt)
    ws.set_row(0, 30)

    # Family group headers
    ws.write(1, 0, "", hdr_fmt)
    ws.write(1, 1, "", hdr_fmt)
    ws.write(1, 2, "", hdr_fmt)
    ws.write(1, 3, "", hdr_fmt)
    col = FIXED_COLS
    for fam_name in fam_names:
        ws.merge_range(1, col, 1, col + COLS_PER_FAM - 1, fam_name, sub_fmt)
        col += COLS_PER_FAM
    ws.set_row(1, 25)

    # Column headers
    ws.write(2, 0, "Tag Label",  hdr_fmt)
    ws.write(2, 1, "Function",   hdr_fmt)
    ws.write(2, 2, "Unit",       hdr_fmt)
    ws.write(2, 3, "EM Module",  hdr_fmt)
    ws.set_column(0, 0, 18)
    ws.set_column(1, 1, 32)
    ws.set_column(2, 2, 14)
    ws.set_column(3, 3, 14)

    sub_cols = ["Excel\nVal", "Exp\nEN", "Exp\nEX", "Exp\nVF",
                "DB\nEN", "DB\nEX", "DB\nVF", "Match", "Action"]
    widths   = [8, 7, 7, 7, 7, 7, 7, 10, 28]
    col = FIXED_COLS
    for fam_name in fam_names:
        for h, w in zip(sub_cols, widths):
            ws.write(2, col, h, hdr_fmt)
            ws.set_column(col, col, w)
            col += 1
    ws.set_row(2, 40)
    ws.freeze_panes(3, FIXED_COLS)

    def bstr(v):
        if isinstance(v, bool):
            return "T" if v else "F"
        return str(v)

    # Data rows
    ri = 3
    for rec in comparison_rows:
        ws.write(ri, 0, rec.get("Tag Label", ""), norm_fmt)
        ws.write(ri, 1, rec.get("Function",   ""), norm_fmt)
        ws.write(ri, 2, rec.get("Unit",        ""), norm_fmt)
        ws.write(ri, 3, rec.get("EM",          ""), norm_fmt)

        col = FIXED_COLS
        for fam_name in fam_names:
            fs = fam_name.replace(" ", "")
            excel_val = str(rec.get(f"{fs}_Excel", "")).upper()
            exp_en    = rec.get(f"{fs}_ExpEN", False)
            exp_ex    = rec.get(f"{fs}_ExpEX", False)
            exp_vf    = rec.get(f"{fs}_ExpVF", "N/A")
            db_en     = rec.get(f"{fs}_DbEN",  False)
            db_ex     = rec.get(f"{fs}_DbEX",  False)
            db_vf     = rec.get(f"{fs}_DbVF",  "N/A")
            match     = rec.get(f"{fs}_Match", "OK")
            action    = rec.get(f"{fs}_Action", "OK")

            # Excel val
            efmt = ok_fmt if excel_val == "X" else (warn_fmt if excel_val == "O" else norm_fmt)
            ws.write(ri, col, excel_val if excel_val not in ("NONE", "") else "-", efmt); col += 1

            # Expected
            ws.write(ri, col, bstr(exp_en), ok_fmt if exp_en else norm_fmt); col += 1
            ws.write(ri, col, bstr(exp_ex), ok_fmt if exp_ex else norm_fmt); col += 1
            ws.write(ri, col, bstr(exp_vf), norm_fmt); col += 1

            # DB effective
            ws.write(ri, col, bstr(db_en),  ok_fmt if db_en else norm_fmt); col += 1
            ws.write(ri, col, bstr(db_ex),  ok_fmt if db_ex else norm_fmt); col += 1
            ws.write(ri, col, bstr(db_vf),  norm_fmt); col += 1

            # Match
            if match == "OK":
                ws.write(ri, col, "OK", ok_fmt)
            elif "NOT" in action:
                ws.write(ri, col, "N/F", warn_fmt)
            else:
                ws.write(ri, col, "DIFF", fail_fmt)
            col += 1

            # Action
            if action == "OK":
                ws.write(ri, col, action, ok_fmt)
            elif "NOT" in action:
                ws.write(ri, col, action, warn_fmt)
            else:
                ws.write(ri, col, action, fail_fmt)
            col += 1

        ri += 1

    # Summary sheet
    ws2 = wb.add_worksheet("Summary")
    ws2.merge_range(0, 0, 0, 4, "Comparison Summary", title_fmt)
    ws2.set_row(0, 30)
    for ci, h in enumerate(["Family", "Total", "OK", "Mismatches", "Not Found"]):
        ws2.write(1, ci, h, hdr_fmt)
        ws2.set_column(ci, ci, 18)

    ri2 = 2
    for fam_name in fam_names:
        fs    = fam_name.replace(" ", "")
        total = len(comparison_rows)
        ok_c  = sum(1 for r in comparison_rows if r.get(f"{fs}_Action") == "OK")
        mm_c  = sum(1 for r in comparison_rows if "UPDATE" in str(r.get(f"{fs}_Action", "")))
        nf_c  = sum(1 for r in comparison_rows if "NOT" in str(r.get(f"{fs}_Action", "")))
        ws2.write(ri2, 0, fam_name, norm_fmt)
        ws2.write(ri2, 1, total,    center_fmt)
        ws2.write(ri2, 2, ok_c,     ok_fmt)
        ws2.write(ri2, 3, mm_c,     fail_fmt if mm_c else ok_fmt)
        ws2.write(ri2, 4, nf_c,     warn_fmt if nf_c else ok_fmt)
        ri2 += 1

    wb.close()
    print(f"  Written: {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Brew DB Instrument Configuration Checker")
    print("=" * 60)

    print(f"\n[1] Reading DB: {DB_PATH}")
    with open(DB_PATH, "r", encoding="utf-8-sig") as f:
        db_text = f.read()
    print(f"    {db_text.count(chr(10))} lines")

    print(f"\n[2] Reading Excel: {EXCEL_PATH}")
    excel_data = parse_excel(EXCEL_PATH)
    for fam, data in excel_data.items():
        print(f"    {fam}: {len(data)} instruments")

    print("\n[3] Parsing DB type defaults...")
    type_defaults = parse_type_defaults(db_text)
    print(f"    {len(type_defaults)} defaults parsed")

    # Debug: show a few key entries
    for key in ["Unit Sep|EM - 400|FIS409-1", "Unit Sep|EM - 100|EMS",
                "Unit Sep|EM - 100|UPS", "Unit Process|EM - 22Y|AV221-1"]:
        if key in type_defaults:
            td = type_defaults[key]
            print(f"    {key.split('|')[-1]}: EN={td['default_enable']} EX={td['default_exist']} VF={td['default_vlv_w_fb']}")

    print("\n[4] Parsing MachineConfig overrides...")
    mc_overrides_all = {}
    for fam_name, fam_cfg in FAMILIES.items():
        mc_idx = fam_cfg["mc_idx"]
        ov = parse_mc_overrides(db_text, mc_idx)
        mc_overrides_all[mc_idx] = ov
        print(f"    MC[{mc_idx}] ({fam_name}): {len(ov)} CW overrides")

    print("\n[5] Building comparison...")
    comparison_rows = build_comparison(excel_data, type_defaults, mc_overrides_all)
    print(f"    {len(comparison_rows)} instruments compared")

    print("\n    === Mismatch Summary ===")
    for fam_name in FAMILIES:
        fs = fam_name.replace(" ", "")
        ok_c = sum(1 for r in comparison_rows if r.get(f"{fs}_Action") == "OK")
        mm   = [r for r in comparison_rows if "UPDATE" in str(r.get(f"{fs}_Action", ""))]
        nf   = [r for r in comparison_rows if "NOT"    in str(r.get(f"{fs}_Action", ""))]
        print(f"    {fam_name}: OK={ok_c}  MISMATCH={len(mm)}  NOT_FOUND={len(nf)}")
        for r in mm:
            act = r[f"{fs}_Action"].encode("ascii", "replace").decode()
            print(f"      >> {r['Tag Label']:20s} {act}")

    print(f"\n[6] Writing comparison report: {OUT_XLSX}")
    write_comparison_xlsx(comparison_rows, OUT_XLSX)

    print(f"\n[7] Generating updated DB: {OUT_DB}")
    updated_db = generate_updated_db(db_text, excel_data, type_defaults, mc_overrides_all)
    with open(OUT_DB, "w", encoding="utf-8-sig", newline="") as f:
        f.write(updated_db)
    print(f"    Written: {OUT_DB}")

    print("\n[8] Verifying updated DB...")
    ov_after = {}
    for fam_name, fam_cfg in FAMILIES.items():
        ov_after[fam_cfg["mc_idx"]] = parse_mc_overrides(updated_db, fam_cfg["mc_idx"])

    mismatch_after = 0
    for fam_name, fam_cfg in FAMILIES.items():
        mc_idx   = fam_cfg["mc_idx"]
        fam_data = excel_data.get(fam_name, {})
        for excel_tag, info in fam_data.items():
            if excel_tag not in TAG_MAP:
                continue
            eff_en, eff_ex, _, _ = compute_effective(type_defaults, ov_after[mc_idx], excel_tag)
            if eff_en != info["enable"] or eff_ex != info["exist"]:
                mismatch_after += 1
                print(f"  STILL MISMATCH [{fam_name}] {excel_tag}: "
                      f"EN={eff_en}(exp {info['enable']}) EX={eff_ex}(exp {info['exist']})")

    if mismatch_after == 0:
        print("    All ENABLE/EXIST values verified correct.")
    else:
        print(f"    WARNING: {mismatch_after} mismatches remain after update!")

    print("\n" + "=" * 60)
    print("COMPLETE")
    print(f"  Report    : {OUT_XLSX}")
    print(f"  Updated DB: {OUT_DB}")
    print("=" * 60)


if __name__ == "__main__":
    main()
