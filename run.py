"""
Universal TIA Portal DB Instrument Configuration Checker
=========================================================
Auto-discovers instrument paths from DB TYPE blocks — NO manual TAG_MAP needed.

For each job defined in jobs.json:
  - Reads the specified .db and .xlsx files
  - Auto-builds the tag map from the DB structure
  - Compares ENABLE / EXIST / VLV_W_FB / NO / FB_OPN_EN / FB_CLS_EN per instrument × family
  - Produces a separate comparison .xlsx and corrected .db per job

CW bit layout (CW_VLV / CW_AN):
  bit  3 = ENABLE         (all STAT types)
  bit  6 = FB_CLS_EN      (CW_VLV only — Enable Close Feedback)
  bit  7 = FB_OPN_EN      (CW_VLV only — Enable Open Feedback)
  bit 12 = NO             (CW_VLV and CW_AN — 0=NC, 1=NO)
  bit 13 = EXIST          (all STAT types)
  bit 15 = VLV_W_FB       (CW_VLV only — Valve Without Feedback)

Excel Type column → NO bit:
  NC  → NO = False  (0=NC, Normally Closed)
  NO  → NO = True   (1=NO, Normally Open)
  other/empty → no override (keep type default)

Usage:
  python run.py                  # run all jobs in jobs.json
  python run.py Brew             # run one job by name
  python run.py --list           # list available jobs without running

Edit jobs.json to add / modify products and families.
"""

import re
import sys
import json
import openpyxl
import xlsxwriter
from collections import defaultdict

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

JOBS_FILE = "jobs.json"
UDT_FILE  = "STAT SYSTEM.udt"

# ── Low-level helpers ──────────────────────────────────────────────────────────

def norm_bool(v):
    return v.strip().lower() in ("true",)

def path_key(unit, em, field):
    return f"{unit}|{em}|{field}"

def extract_tuple_content(text, start_pos):
    """Bracket-depth walker — returns content between matching parens."""
    depth = 0
    for i in range(start_pos, len(text)):
        if text[i] == "(":
            depth += 1
        elif text[i] == ")":
            depth -= 1
            if depth == 0:
                return text[start_pos + 1 : i]
    return text[start_pos + 1:]

def parse_cw_bits(cw_content):
    """Parse 16-element CW tuple content → list of booleans."""
    bits = []
    i = 0
    s = cw_content
    while i < len(s):
        c = s[i]
        if c in " \t\r\n,":
            i += 1
        elif c == "(":
            bits.append(False)
            i = s.find(")", i) + 1
        elif c in "TtFf":
            j = i
            while j < len(s) and s[j].isalpha():
                j += 1
            bits.append(s[i:j].upper() == "TRUE")
            i = j
        else:
            j = i
            while j < len(s) and s[j] not in ",)":
                j += 1
            bits.append(False)
            i = j
    return bits

def make_db_line(mc_idx, unit, em, field, prop, value):
    bool_str = "True" if value else "False"
    def q(s):
        # Quote if the name contains any character that is not plain alphanumeric or underscore
        return f'"{s}"' if re.search(r'[^A-Za-z0-9_]', s) else s
    return f'   MachineConfig[{mc_idx}]."{unit}"."{em}".{q(field)}.CW.{prop} := {bool_str};\r\n'


# ── Auto-discover TAG_MAP from DB ──────────────────────────────────────────────

STAT_TYPES = {"STAT VLV", "STAT DI", "STAT DO", "STAT AI", "STAT AO", "STAT MTR", "STAT VFD"}


def build_em_unit_from_udt(udt_text):
    """
    Parse STAT SYSTEM.udt to build the authoritative EM → Unit mapping.

    Reads the two-level hierarchy defined in the UDT:
      TYPE "STAT SYSTEM"  →  maps each unit field name to its STAT_xxx type
      TYPE "STAT Sep/Process/Dch/Options/Comm"  →  lists the EM modules per unit

    Returns: {em_name: unit_name}
    e.g. {"EM - 100": "Unit Sep", "EM - 208": "Options", "EM - PRU1": "Options", ...}

    This is independent of any MachineConfig[N] instance data, so every EM is
    mapped correctly even if it has no override lines in the DB file.
    """
    type_pat  = re.compile(r'TYPE\s+"([^"]+)".*?END_TYPE', re.DOTALL)
    field_pat = re.compile(
        r'(?:"([^"]+)"|([A-Za-z][A-Za-z0-9_]*))'   # grp1=quoted, grp2=unquoted field name
        r'\s*(?:\{[^}]*\})?\s*'                     # optional { attributes } block
        r':\s*"([^"]+)"',                            # : "TypeName"
        re.DOTALL
    )

    # Pass 1: STAT SYSTEM block → {stat_type_name: unit_name}
    # e.g. {"STAT Sep": "Unit Sep", "STAT Options": "Options", ...}
    stat_type_to_unit = {}
    for type_m in type_pat.finditer(udt_text):
        if type_m.group(1) != "STAT SYSTEM":
            continue
        for fm in field_pat.finditer(type_m.group(0)):
            unit_field = fm.group(1) or fm.group(2)
            type_name  = fm.group(3)
            if type_name.startswith("STAT "):   # skip Machine : "MachineConfig"
                stat_type_to_unit[type_name] = unit_field
        break   # only one STAT SYSTEM block

    # Pass 2: for each STAT_xxx block, collect EM field names → unit_name
    em_unit = {}
    for type_m in type_pat.finditer(udt_text):
        type_name = type_m.group(1)
        if type_name not in stat_type_to_unit:
            continue
        unit = stat_type_to_unit[type_name]
        for fm in field_pat.finditer(type_m.group(0)):
            em_name = fm.group(1) or fm.group(2)
            if em_name.startswith("EM"):
                em_unit[em_name] = unit

    return em_unit


def build_tag_map(db_text, em_unit):
    """
    Auto-build {field_name: (unit, em, field_name, is_vlv)} from DB TYPE blocks.

    em_unit: {em_name: unit_name} — derived from STAT SYSTEM.udt via
             build_em_unit_from_udt(). Replaces the old BEGIN-section scan so
             every EM is mapped correctly regardless of whether it has any
             MachineConfig[N] override lines in this specific DB file.

    Scans TYPE "EM-xxx" blocks for STAT-typed instrument fields.
    """
    tag_map   = {}
    type_pat  = re.compile(r'TYPE\s+"([^"]+)".*?END_TYPE', re.DOTALL)
    # Matches both  "Quoted Name" { attrs } : "STAT XX" :=
    #           and  UnquotedName { attrs }  : "STAT XX" :=
    field_pat = re.compile(
        r'(?:"([^"]+)"|([A-Za-z0-9_\-\+\./]+))'   # grp1=quoted, grp2=unquoted
        r'(?:\s*\{[^}]*\})?'                        # optional attribute block
        r'\s*:\s*"(STAT [A-Z]+)"\s*:=',
        re.DOTALL
    )

    for type_m in type_pat.finditer(db_text):
        em = type_m.group(1)
        if not em.startswith("EM"):
            continue
        unit = em_unit.get(em, "Unknown")
        body = type_m.group(0)
        for fm in field_pat.finditer(body):
            field     = (fm.group(1) or fm.group(2)).strip()
            stat_type = fm.group(3)
            if stat_type not in STAT_TYPES:
                continue
            is_vlv = (stat_type == "STAT VLV")
            if field not in tag_map:          # first EM occurrence wins
                tag_map[field] = (unit, em, field, is_vlv)

    return tag_map


# ── Step 1: Parse Excel ────────────────────────────────────────────────────────

def parse_excel(path, families_cfg, tag_map):
    """
    families_cfg: {family_name: mc_idx}  (from jobs.json)

    Returns: {family_name: {tag_label: {enable, exist, vlv_w_fb, fb_opn_en, fb_cls_en,
                                        no_val, is_vlv, ...}}}

    Type column logic (col header = "Type"):
      NC   → no_val = False  (Normally Closed output — NO bit = False, 0=NC)
      NO   → no_val = True   (Normally Open output  — NO bit = True,  1=NO)
      other/empty → no_val = None  (no override; keep type default)

    VLV_W_FB / FB_OPN_EN / FB_CLS_EN (AV valves only):
      FB OPN = x or o  →  fb_opn_en = True,  VLV_W_FB = True
      FB CLS = x or o  →  fb_cls_en = True,  VLV_W_FB = True
      both empty/spare →  VLV_W_FB = False

    Option Module Rule (unit = "Options" in DB):
      ENABLE = False, EXIST = False unconditionally.
      VLV_W_FB / FB_OPN_EN / FB_CLS_EN / NO follow normal logic.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Auto-detect family columns, option column, and type column from header rows
    family_cols = {}    # {col_idx (1-based): family_name}
    option_col  = None
    type_col    = None
    header_row  = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                if v in families_cfg:
                    family_cols[cell.column] = v
                    header_row = cell.row
                elif v.lower() == "option":
                    option_col = cell.column
                elif v.lower() == "type":
                    type_col = cell.column

    if not family_cols:
        raise ValueError(
            f"No matching family columns found in {path}.\n"
            f"  Looking for: {list(families_cfg.keys())}"
        )

    # Tags under "Options" unit in the DB are always ENABLE=False, EXIST=False
    option_tags = {tag for tag, info in tag_map.items() if info[0] == "Options"}

    def type_to_no(type_val):
        """Convert Excel Type column value to NO bit. None = no override.
        DB comment: NO : Bool  // 0=NC, 1=NO
          NC  → NO = False  (0 = Normally Closed)
          NO  → NO = True   (1 = Normally Open)
          other/empty → None (keep type default)
        """
        if type_val is None:
            return None
        v = str(type_val).strip().upper()
        if v == "NC":
            return False
        if v == "NO":
            return True
        return None   # 4..20MA or other analog types — no NO override

    # Read data rows
    raw_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        tag_label = row[1].value
        func      = row[2].value if row[2].value else ""
        if not tag_label:
            continue
        tag_label = str(tag_label).strip()
        func      = str(func).strip()
        if tag_label.upper() in ("SPARE", "DUMMY", "") or func.upper() in ("SPARE", "DUMMY"):
            continue

        fam_vals = {}
        for col_idx, fam_name in family_cols.items():
            cell    = row[col_idx - 1]
            comment = None
            if cell.comment:
                lines = [ln.strip() for ln in cell.comment.text.splitlines() if ln.strip()]
                for ln in reversed(lines):
                    if not any(x in ln.lower() for x in
                               ("http", "version", "learn", "excel", "comment:", "threaded")):
                        comment = ln
                        break
            fam_vals[fam_name] = {"val": cell.value, "comment": comment}

        opt_desc = ""
        if option_col:
            opt_desc = str(row[option_col - 1].value or "").strip().upper()

        no_val = None
        if type_col:
            no_val = type_to_no(row[type_col - 1].value)

        raw_rows.append((tag_label, func, fam_vals, opt_desc, no_val))

    # Aggregate: AV valves (FB OPN / FB CLS / ACT rows) vs everything else
    valve_rows     = defaultdict(lambda: {"fb_opn": {}, "fb_cls": {}, "act": {}, "no_val": None})
    non_valve_rows = {}
    AV_FUNCS       = {"FB OPN", "FB CLS", "ACT"}

    for tag_label, func, fam_vals, _opt, no_val in raw_rows:
        func_up = func.upper().strip()
        if tag_label.upper().startswith("AV") and func_up in AV_FUNCS:
            key = {"FB OPN": "fb_opn", "FB CLS": "fb_cls", "ACT": "act"}[func_up]
            valve_rows[tag_label][key] = fam_vals
            if no_val is not None:
                valve_rows[tag_label]["no_val"] = no_val
        else:
            if tag_label not in non_valve_rows:
                non_valve_rows[tag_label] = {"fam_vals": {}, "func": func, "no_val": no_val}
            for fam, fv in fam_vals.items():
                existing = non_valve_rows[tag_label]["fam_vals"]
                if fam not in existing or (existing[fam]["val"] is None and fv["val"] is not None):
                    existing[fam] = fv
            if no_val is not None:
                non_valve_rows[tag_label]["no_val"] = no_val

    def xlval(raw_val, is_option=False):
        """Return (enable, exist) from Excel cell value."""
        if is_option:
            return False, False
        if raw_val is None:
            return False, False
        v = str(raw_val).strip().lower()
        if v == "x":   return True, True
        elif v == "o": return True, False
        return False, False

    def fb_active(fam_vals, fam_name):
        """True if the FB OPN or FB CLS cell is x or o for this family."""
        fv = fam_vals.get(fam_name, {}).get("val")
        if fv is None:
            return False
        return str(fv).strip().lower() in ("x", "o")

    families_data = {fam: {} for fam in family_cols.values()}

    # AV valves
    for tag_label, rows in valve_rows.items():
        tag_no_val = rows["no_val"]
        for fam_name in family_cols.values():
            is_opt  = tag_label in option_tags
            act_fv  = rows["act"].get(fam_name, {"val": None, "comment": None})
            enable, exist = xlval(act_fv.get("val"), is_opt)

            fb_opn_active = fb_active(rows["fb_opn"], fam_name)
            fb_cls_active = fb_active(rows["fb_cls"], fam_name)
            fb_opn_en  = fb_opn_active
            fb_cls_en  = fb_cls_active
            vlv_w_fb   = fb_opn_active or fb_cls_active

            families_data[fam_name][tag_label] = {
                "enable":     enable,
                "exist":      exist,
                "vlv_w_fb":   vlv_w_fb,
                "fb_opn_en":  fb_opn_en,
                "fb_cls_en":  fb_cls_en,
                "no_val":     tag_no_val,
                "resolved_tag": act_fv.get("comment") or tag_label,
                "label":      tag_label,
                "function":   "AV Valve",
                "is_vlv":     True,
                "raw_act":    act_fv.get("val"),
                "raw_fb_opn": rows["fb_opn"].get(fam_name, {}).get("val"),
                "raw_fb_cls": rows["fb_cls"].get(fam_name, {}).get("val"),
            }

    # Non-valve instruments
    for tag_label, info in non_valve_rows.items():
        for fam_name in family_cols.values():
            is_opt  = tag_label in option_tags
            fv      = info["fam_vals"].get(fam_name, {"val": None, "comment": None})
            raw_val = fv.get("val")
            enable, exist = xlval(raw_val, is_opt)
            families_data[fam_name][tag_label] = {
                "enable":    enable,
                "exist":     exist,
                "vlv_w_fb":  False,
                "fb_opn_en": False,
                "fb_cls_en": False,
                "no_val":    info.get("no_val"),
                "resolved_tag": fv.get("comment") or tag_label,
                "label":     tag_label,
                "function":  info["func"],
                "is_vlv":    False,
                "raw_val":   raw_val,
            }

    return families_data


# ── Step 2: Parse DB type defaults ────────────────────────────────────────────

def parse_type_defaults(db_text, tag_map):
    """
    Parse TYPE...END_TYPE blocks using em→unit from tag_map.

    CW bit positions extracted:
      bit  3 = ENABLE       (all STAT types)
      bit  6 = FB_CLS_EN    (CW_VLV / STAT VLV only)
      bit  7 = FB_OPN_EN    (CW_VLV / STAT VLV only)
      bit 12 = NO           (CW_VLV and CW_AN)
      bit 13 = EXIST        (all STAT types)
      bit 15 = VLV_W_FB     (CW_VLV / STAT VLV only)

    Returns: {path_key: {default_enable, default_exist, default_vlv_w_fb,
                          default_no, default_fb_opn_en, default_fb_cls_en, ...}}
    """
    em_unit   = {info[1]: info[0] for info in tag_map.values()}
    defaults  = {}
    type_pat  = re.compile(r'TYPE\s+"([^"]+)".*?END_TYPE', re.DOTALL)
    field_pat = re.compile(
        r'(?:"([^"]+)"|([A-Za-z0-9_\-\+\./]+))'
        r'(?:\s*\{[^}]*\})?'
        r'\s*:\s*"(STAT [A-Z]+)"\s*:=\s*',
        re.DOTALL
    )

    for type_m in type_pat.finditer(db_text):
        em = type_m.group(1)
        if em not in em_unit:
            continue
        unit = em_unit[em]
        body = type_m.group(0)

        for fm in field_pat.finditer(body):
            field_raw = (fm.group(1) or fm.group(2)).strip()
            stat_type = fm.group(3).strip()
            if stat_type not in STAT_TYPES:
                continue
            is_vlv = (stat_type == "STAT VLV")

            p = fm.end()
            while p < len(body) and body[p] in " \t\r\n":
                p += 1
            if p >= len(body) or body[p] != "(":
                continue

            outer = extract_tuple_content(body, p)
            q1 = outer.find("'")
            if q1 == -1: continue
            q2 = outer.find("'", q1 + 1)
            if q2 == -1: continue
            cp = outer.find("(", q2 + 1)
            if cp == -1: continue
            bits = parse_cw_bits(extract_tuple_content(outer, cp))

            pk = path_key(unit, em, field_raw)
            defaults[pk] = {
                "default_enable":    bits[3]  if len(bits) > 3  else False,
                "default_fb_cls_en": bits[6]  if (is_vlv and len(bits) > 6)  else False,
                "default_fb_opn_en": bits[7]  if (is_vlv and len(bits) > 7)  else False,
                "default_no":        bits[12] if len(bits) > 12 else False,
                "default_exist":     bits[13] if len(bits) > 13 else False,
                "default_vlv_w_fb":  bits[15] if (is_vlv and len(bits) > 15) else False,
                "unit": unit, "em": em, "field": field_raw, "is_vlv": is_vlv,
            }

    return defaults


# ── Step 3: Parse MachineConfig[N] overrides ──────────────────────────────────

# All CW properties that can appear as overrides in the BEGIN section
CW_PROPS = {"ENABLE", "EXIST", "VLV_W_FB", "NO", "FB_OPN_EN", "FB_CLS_EN"}

def parse_mc_overrides(db_text, mc_idx):
    overrides = defaultdict(dict)
    pat = re.compile(
        r'MachineConfig\[' + str(mc_idx) + r'\]\.'
        r'(?:"([^"]+)"|([A-Za-z][A-Za-z0-9_]*))\."([^"]+)"\.'  # unit (q/unq) + em (quoted)
        r'(?:"([^"]+)"|([A-Za-z0-9_\-]+))\.'                    # field (q/unq)
        r'CW\.(ENABLE|EXIST|VLV_W_FB|NO|FB_OPN_EN|FB_CLS_EN)\s*:=\s*(True|true|False|false)\s*;',
        re.IGNORECASE
    )
    for m in pat.finditer(db_text):
        unit  = m.group(1) or m.group(2)
        em    = m.group(3)
        field = m.group(4) if m.group(4) else m.group(5)
        prop  = m.group(6).upper()
        val   = norm_bool(m.group(7))
        pk    = path_key(unit, em, field)
        if prop == "ENABLE":      overrides[pk]["enable"]      = val
        elif prop == "EXIST":     overrides[pk]["exist"]       = val
        elif prop == "VLV_W_FB":  overrides[pk]["vlv_w_fb"]   = val
        elif prop == "NO":        overrides[pk]["no"]          = val
        elif prop == "FB_OPN_EN": overrides[pk]["fb_opn_en"]  = val
        elif prop == "FB_CLS_EN": overrides[pk]["fb_cls_en"]  = val
    return dict(overrides)


# ── Step 4: Compute effective value ───────────────────────────────────────────

def compute_effective(type_defaults, mc_overrides, excel_tag, tag_map):
    if excel_tag not in tag_map:
        return False, False, False, False, False, False, False
    unit, em, field, is_vlv = tag_map[excel_tag]
    pk  = path_key(unit, em, field)
    td  = type_defaults.get(pk, {})
    ov  = mc_overrides.get(pk, {})
    found = (pk in type_defaults) or (pk in mc_overrides)
    return (
        ov.get("enable",      td.get("default_enable",    False)),
        ov.get("exist",       td.get("default_exist",     False)),
        ov.get("vlv_w_fb",    td.get("default_vlv_w_fb",  False)),
        ov.get("no",          td.get("default_no",        False)),
        ov.get("fb_opn_en",   td.get("default_fb_opn_en", False)),
        ov.get("fb_cls_en",   td.get("default_fb_cls_en", False)),
        found,
    )


# ── Step 4b: Resolve DB field name from Excel comment + tag label ─────────────

def _resolve_db_field(resolved_raw, tag_label, tag_map, tag_map_lower):
    """Priority-based DB field resolution.

    resolved_raw : inst.get("resolved_tag", tag_label)  — may be a cell comment
    tag_label    : the Excel Col B label (always the key in fam_data dict)

    Priority:
      1. resolved_raw exact in tag_map   → Case B exact  (e.g. TT730a)
      2. tag_label    exact in tag_map   → Case A        (e.g. PCV22X-1)
      3. resolved_raw case-insensitive   → Case B casing (e.g. TT731A → TT731a)
      4. Strip "Word: " prefix, retry 1+3 (e.g. Sensor: TT733 → TT733)
      5. Return resolved_raw as-is       → genuinely absent → NOT IN DB MAP
    """
    if resolved_raw in tag_map:
        return resolved_raw
    if tag_label in tag_map:
        return tag_label
    ci = tag_map_lower.get(resolved_raw.lower())
    if ci:
        return ci
    stripped = re.sub(r'^\w+:\s+', '', resolved_raw).strip()
    if stripped != resolved_raw:
        if stripped in tag_map:
            return stripped
        ci2 = tag_map_lower.get(stripped.lower())
        if ci2:
            return ci2
    return resolved_raw


# ── Step 5: Build comparison rows ─────────────────────────────────────────────

def build_comparison(excel_data, type_defaults, mc_overrides_all, tag_map, families_cfg):
    # Collect ordered unique tag labels across all families
    seen = {}
    for fam_data in excel_data.values():
        for tag_label, info in fam_data.items():
            if tag_label not in seen:
                seen[tag_label] = info.get("function", "")

    tag_map_lower = {k.lower(): k for k in tag_map}
    rows = []
    for tag_label, function in seen.items():
        # ── Resolve the base tag_map entry for fixed columns (Unit, EM, is_vlv).
        # Case A: Excel label == DB field name → direct hit.
        # Case B: Excel label is a display placeholder (e.g. ST74X); per-family cell
        #         comment holds the actual DB field name (e.g. ST741/ST740).
        #         Scan families for the first resolvable entry.
        base_entry = tag_map.get(tag_label)
        if base_entry is None:
            for _fn in families_cfg:
                _inst = excel_data.get(_fn, {}).get(tag_label)
                if _inst:
                    _rt = _resolve_db_field(
                        _inst.get("resolved_tag", tag_label), tag_label,
                        tag_map, tag_map_lower)
                    if _rt in tag_map:
                        base_entry = tag_map[_rt]
                        break

        row = {
            "Tag Label": tag_label,
            "Function":  function,
            "Unit": base_entry[0] if base_entry else "N/A",
            "EM":   base_entry[1] if base_entry else "N/A",
        }

        for fam_name, mc_idx in families_cfg.items():
            mc_overrides = mc_overrides_all.get(mc_idx, {})
            fs           = fam_name.replace(" ", "")
            fam_data     = excel_data.get(fam_name, {})

            if tag_label in fam_data:
                inst     = fam_data[tag_label]
                # Priority resolution: try comment exact → tag_label exact →
                # case-insensitive → strip prefix → as-is (see _resolve_db_field)
                resolved = _resolve_db_field(
                    inst.get("resolved_tag", tag_label), tag_label,
                    tag_map, tag_map_lower)
                in_map   = resolved in tag_map
                is_vlv   = tag_map[resolved][3] if in_map else (base_entry[3] if base_entry else False)
                exp_en   = inst["enable"]
                exp_ex   = inst["exist"]
                exp_vf   = inst["vlv_w_fb"]  if is_vlv else False
                exp_no   = inst["no_val"]     # None = no override expected
                exp_fbopn = inst["fb_opn_en"] if is_vlv else False
                exp_fbcls = inst["fb_cls_en"] if is_vlv else False
                raw_disp  = (
                    str(inst.get("raw_act", "")).upper() or "empty"
                    if inst.get("is_vlv") else
                    str(inst.get("raw_val", "")).upper() or "empty"
                )
            else:
                resolved  = tag_label
                in_map    = resolved in tag_map
                is_vlv    = base_entry[3] if base_entry else False
                exp_en, exp_ex, exp_vf = False, False, False
                exp_no, exp_fbopn, exp_fbcls = None, False, False
                raw_disp = "N/A"

            if in_map:
                eff_en, eff_ex, eff_vf, eff_no, eff_fbopn, eff_fbcls, found = \
                    compute_effective(type_defaults, mc_overrides, resolved, tag_map)
            else:
                eff_en, eff_ex, eff_vf, eff_no, eff_fbopn, eff_fbcls, found = \
                    False, False, False, False, False, False, False

            en_m    = (exp_en  == eff_en)
            ex_m    = (exp_ex  == eff_ex)
            vf_m    = (exp_vf  == eff_vf)   if (in_map and is_vlv) else True
            no_m    = (exp_no  == eff_no)   if (in_map and exp_no is not None) else True
            fbopn_m = (exp_fbopn == eff_fbopn) if (in_map and is_vlv) else True
            fbcls_m = (exp_fbcls == eff_fbcls) if (in_map and is_vlv) else True

            if not in_map:
                action = "NOT IN DB MAP"
            elif not found:
                action = "NOT FOUND IN DB"
            elif not en_m or not ex_m or not vf_m or not no_m or not fbopn_m or not fbcls_m:
                parts = []
                if not en_m:    parts.append(f"ENABLE: DB={eff_en}->{exp_en}")
                if not ex_m:    parts.append(f"EXIST: DB={eff_ex}->{exp_ex}")
                if is_vlv and not vf_m:
                    parts.append(f"VLV_W_FB: DB={eff_vf}->{exp_vf}")
                if exp_no is not None and not no_m:
                    parts.append(f"NO: DB={eff_no}->{exp_no}")
                if is_vlv and not fbopn_m:
                    parts.append(f"FB_OPN_EN: DB={eff_fbopn}->{exp_fbopn}")
                if is_vlv and not fbcls_m:
                    parts.append(f"FB_CLS_EN: DB={eff_fbcls}->{exp_fbcls}")
                action = ("UPDATE: " + ", ".join(parts)) if parts else "OK"
            else:
                action = "OK"

            all_match = en_m and ex_m and vf_m and no_m and fbopn_m and fbcls_m
            row.update({
                f"{fs}_Excel":    raw_disp,
                f"{fs}_Type":     "NO" if exp_no is True else ("NC" if exp_no is False else "-"),
                f"{fs}_ExpEN":    exp_en,
                f"{fs}_ExpEX":    exp_ex,
                f"{fs}_ExpVF":    exp_vf      if is_vlv else "N/A",
                f"{fs}_ExpNO":    exp_no      if exp_no is not None else "-",
                f"{fs}_ExpFBOPN": exp_fbopn   if is_vlv else "N/A",
                f"{fs}_ExpFBCLS": exp_fbcls   if is_vlv else "N/A",
                f"{fs}_DbEN":     eff_en,
                f"{fs}_DbEX":     eff_ex,
                f"{fs}_DbVF":     eff_vf      if is_vlv else "N/A",
                f"{fs}_DbNO":     eff_no,
                f"{fs}_DbFBOPN":  eff_fbopn   if is_vlv else "N/A",
                f"{fs}_DbFBCLS":  eff_fbcls   if is_vlv else "N/A",
                f"{fs}_Match":    "OK" if all_match else "MISMATCH",
                f"{fs}_Action":   action,
            })
        rows.append(row)
    return rows


# ── Step 6: Generate updated DB ───────────────────────────────────────────────

def generate_updated_db(db_text, excel_data, type_defaults, mc_overrides_all,
                        tag_map, families_cfg):
    lines = db_text.splitlines(keepends=True)

    begin_idx = end_idx = None
    for i, line in enumerate(lines):
        s = line.strip()
        if s == "BEGIN":          begin_idx = i
        if s == "END_DATA_BLOCK": end_idx   = i

    if begin_idx is None or end_idx is None:
        print("WARNING: Could not find BEGIN/END_DATA_BLOCK — DB unchanged.")
        return db_text

    # Patterns to strip old CW override lines per mc_idx
    # Unit may be quoted ("Unit Sep") or unquoted (Options, Communications)
    remove_pats = {
        mc_idx: re.compile(
            r'MachineConfig\[' + str(mc_idx) + r'\]\.'
            r'(?:"[^"]+"|[A-Za-z][A-Za-z0-9_]*)\."[^"]+"\.'
            r'(?:"[^"]+"|[A-Za-z0-9_\-]+)\.'
            r'CW\.(ENABLE|EXIST|VLV_W_FB|NO|FB_OPN_EN|FB_CLS_EN)\s*:=',
            re.IGNORECASE
        )
        for mc_idx in families_cfg.values()
    }

    # Build fresh minimal-patch CW lines per mc_idx
    tag_map_lower = {k.lower(): k for k in tag_map}
    new_cw = {mc_idx: [] for mc_idx in families_cfg.values()}
    for fam_name, mc_idx in families_cfg.items():
        for tag_label, inst in excel_data.get(fam_name, {}).items():
            resolved = _resolve_db_field(
                inst.get("resolved_tag", tag_label), tag_label,
                tag_map, tag_map_lower)
            if resolved not in tag_map:
                continue
            unit, em, field, is_vlv = tag_map[resolved]
            pk  = path_key(unit, em, field)
            td  = type_defaults.get(pk, {})

            def_en     = td.get("default_enable",    False)
            def_ex     = td.get("default_exist",     False)
            def_vf     = td.get("default_vlv_w_fb",  False)
            def_no     = td.get("default_no",        False)
            def_fbopn  = td.get("default_fb_opn_en", False)
            def_fbcls  = td.get("default_fb_cls_en", False)

            exp_en    = inst["enable"]
            exp_ex    = inst["exist"]
            exp_vf    = inst["vlv_w_fb"]  if is_vlv else False
            exp_no    = inst["no_val"]    # None = no override needed
            exp_fbopn = inst["fb_opn_en"] if is_vlv else False
            exp_fbcls = inst["fb_cls_en"] if is_vlv else False

            if exp_en != def_en:
                new_cw[mc_idx].append(make_db_line(mc_idx, unit, em, field, "ENABLE",    exp_en))
            if exp_ex != def_ex:
                new_cw[mc_idx].append(make_db_line(mc_idx, unit, em, field, "EXIST",     exp_ex))
            if is_vlv and exp_vf != def_vf:
                new_cw[mc_idx].append(make_db_line(mc_idx, unit, em, field, "VLV_W_FB",  exp_vf))
            if exp_no is not None and exp_no != def_no:
                new_cw[mc_idx].append(make_db_line(mc_idx, unit, em, field, "NO",        exp_no))
            if is_vlv and exp_fbopn != def_fbopn:
                new_cw[mc_idx].append(make_db_line(mc_idx, unit, em, field, "FB_OPN_EN", exp_fbopn))
            if is_vlv and exp_fbcls != def_fbcls:
                new_cw[mc_idx].append(make_db_line(mc_idx, unit, em, field, "FB_CLS_EN", exp_fbcls))

    # Pass 1: strip old CW lines, track last line position per mc_idx
    cleaned  = []
    last_pos = {mc_idx: -1 for mc_idx in families_cfg.values()}
    for line in lines[begin_idx:end_idx]:
        if any(pat.search(line) for pat in remove_pats.values()):
            continue
        for mc_idx in families_cfg.values():
            if f"MachineConfig[{mc_idx}]" in line:
                last_pos[mc_idx] = len(cleaned)
        cleaned.append(line)

    # Pass 2: reinsert new CW blocks after each mc_idx's last existing line
    inserted = set()
    body = []
    for i, line in enumerate(cleaned):
        body.append(line)
        for mc_idx in sorted(families_cfg.values()):
            if mc_idx not in inserted and i == last_pos[mc_idx]:
                body.append(f'   // === CW auto-patched for MachineConfig[{mc_idx}] ===\r\n')
                body.extend(new_cw[mc_idx])
                inserted.add(mc_idx)

    # Edge case: mc_idx with no existing lines → append before END_DATA_BLOCK
    for mc_idx in sorted(families_cfg.values()):
        if mc_idx not in inserted and new_cw[mc_idx]:
            body.append(f'   // === CW auto-patched for MachineConfig[{mc_idx}] ===\r\n')
            body.extend(new_cw[mc_idx])

    return "".join(lines[:begin_idx]) + "".join(body) + "".join(lines[end_idx:])


# ── Step 7: Verify updated DB ─────────────────────────────────────────────────

def verify_updated_db(db_text, excel_data, type_defaults, tag_map, families_cfg):
    errors = 0
    tag_map_lower    = {k.lower(): k for k in tag_map}
    mc_overrides_new = {mc_idx: parse_mc_overrides(db_text, mc_idx)
                        for mc_idx in families_cfg.values()}
    for fam_name, mc_idx in families_cfg.items():
        ov = mc_overrides_new.get(mc_idx, {})
        for tag_label, inst in excel_data.get(fam_name, {}).items():
            resolved = _resolve_db_field(
                inst.get("resolved_tag", tag_label), tag_label,
                tag_map, tag_map_lower)
            if resolved not in tag_map:
                continue
            unit, em, field, _ = tag_map[resolved]
            pk     = path_key(unit, em, field)
            td     = type_defaults.get(pk, {})
            eff_en = ov.get(pk, {}).get("enable", td.get("default_enable", False))
            eff_ex = ov.get(pk, {}).get("exist",  td.get("default_exist",  False))
            if eff_en != inst["enable"] or eff_ex != inst["exist"]:
                print(f"  VERIFY FAIL [{fam_name}] {tag_label}: "
                      f"EN={eff_en}(exp {inst['enable']}) "
                      f"EX={eff_ex}(exp {inst['exist']})")
                errors += 1
    return errors


# ── Step 8: Write comparison Excel ────────────────────────────────────────────

def write_comparison_xlsx(comparison_rows, out_path, families_cfg, job_name):
    wb  = xlsxwriter.Workbook(out_path)
    ws  = wb.add_worksheet("Instrument Comparison")

    hdr  = wb.add_format({"bold": True, "bg_color": "#003366", "font_color": "white",
                           "border": 1, "align": "center", "valign": "vcenter",
                           "text_wrap": True})
    ok   = wb.add_format({"bg_color": "#C6EFCE", "border": 1, "align": "center"})
    fail = wb.add_format({"bg_color": "#FFC7CE", "border": 1, "align": "center"})
    warn = wb.add_format({"bg_color": "#FFEB9C", "border": 1, "align": "center"})
    norm = wb.add_format({"border": 1})
    titl = wb.add_format({"bold": True, "font_size": 14, "bg_color": "#1F4E79",
                           "font_color": "white", "align": "center", "valign": "vcenter"})
    subh = wb.add_format({"bold": True, "bg_color": "#2E75B6", "font_color": "white",
                           "border": 1, "align": "center"})

    fam_names    = list(families_cfg.keys())
    COLS_PER_FAM = 16   # Excel | Type | ExpEN | ExpEX | ExpVF | ExpNO | ExpFBOPN | ExpFBCLS |
                        #        DbEN  | DbEX  | DbVF  | DbNO  | DbFBOPN | DbFBCLS | Match | Action
    FIXED        = 4
    total_cols   = FIXED + len(fam_names) * COLS_PER_FAM

    ws.merge_range(0, 0, 0, total_cols - 1,
                   f"TIA Portal DB Instrument Status Comparison — {job_name}", titl)
    ws.set_row(0, 30)

    for c in range(FIXED):
        ws.write(1, c, "", hdr)
    col = FIXED
    for fn in fam_names:
        ws.merge_range(1, col, 1, col + COLS_PER_FAM - 1, fn, subh)
        col += COLS_PER_FAM
    ws.set_row(1, 25)

    ws.write(2, 0, "Tag Label", hdr);  ws.set_column(0, 0, 18)
    ws.write(2, 1, "Function",  hdr);  ws.set_column(1, 1, 32)
    ws.write(2, 2, "Unit",      hdr);  ws.set_column(2, 2, 14)
    ws.write(2, 3, "EM Module", hdr);  ws.set_column(3, 3, 14)

    sub_cols = ["Excel\nVal", "Type\n(NC/NO)",
                "Exp\nEN",  "Exp\nEX",  "Exp\nVF",  "Exp\nNO",  "Exp\nFBOPN_EN", "Exp\nFBCLS_EN",
                "DB\nEN",   "DB\nEX",   "DB\nVF",   "DB\nNO",   "DB\nFBOPN_EN",  "DB\nFBCLS_EN",
                "Match",    "Action"]
    widths   = [8, 8,  7, 7, 7, 7, 8, 8,  7, 7, 7, 7, 8, 8,  10, 32]
    col = FIXED
    for fn in fam_names:
        for h, w in zip(sub_cols, widths):
            ws.write(2, col, h, hdr)
            ws.set_column(col, col, w)
            col += 1
    ws.set_row(2, 40)
    ws.freeze_panes(3, FIXED)

    def b(v):
        return ("T" if v else "F") if isinstance(v, bool) else str(v)

    for ri, rec in enumerate(comparison_rows, start=3):
        ws.write(ri, 0, rec.get("Tag Label", ""), norm)
        ws.write(ri, 1, rec.get("Function",   ""), norm)
        ws.write(ri, 2, rec.get("Unit",        ""), norm)
        ws.write(ri, 3, rec.get("EM",          ""), norm)
        col = FIXED
        for fn in fam_names:
            fs     = fn.replace(" ", "")
            xval   = str(rec.get(f"{fs}_Excel", "")).upper()
            typ    = str(rec.get(f"{fs}_Type",  "-"))
            exp_en = rec.get(f"{fs}_ExpEN",    False)
            exp_ex = rec.get(f"{fs}_ExpEX",    False)
            exp_vf = rec.get(f"{fs}_ExpVF",    "N/A")
            exp_no = rec.get(f"{fs}_ExpNO",    "-")
            exp_fbopn = rec.get(f"{fs}_ExpFBOPN", "N/A")
            exp_fbcls = rec.get(f"{fs}_ExpFBCLS", "N/A")
            db_en  = rec.get(f"{fs}_DbEN",     False)
            db_ex  = rec.get(f"{fs}_DbEX",     False)
            db_vf  = rec.get(f"{fs}_DbVF",     "N/A")
            db_no  = rec.get(f"{fs}_DbNO",     False)
            db_fbopn = rec.get(f"{fs}_DbFBOPN","N/A")
            db_fbcls = rec.get(f"{fs}_DbFBCLS","N/A")
            match  = rec.get(f"{fs}_Match",    "OK")
            action = rec.get(f"{fs}_Action",   "OK")

            efmt = ok if xval == "X" else (warn if xval == "O" else norm)
            ws.write(ri, col, xval if xval not in ("NONE","") else "-", efmt); col += 1
            ws.write(ri, col, typ,  norm); col += 1
            ws.write(ri, col, b(exp_en),   ok if exp_en else norm); col += 1
            ws.write(ri, col, b(exp_ex),   ok if exp_ex else norm); col += 1
            ws.write(ri, col, b(exp_vf),   norm); col += 1
            ws.write(ri, col, b(exp_no),   norm); col += 1
            ws.write(ri, col, b(exp_fbopn),norm); col += 1
            ws.write(ri, col, b(exp_fbcls),norm); col += 1
            ws.write(ri, col, b(db_en),    ok if db_en  else norm); col += 1
            ws.write(ri, col, b(db_ex),    ok if db_ex  else norm); col += 1
            ws.write(ri, col, b(db_vf),    norm); col += 1
            ws.write(ri, col, b(db_no),    norm); col += 1
            ws.write(ri, col, b(db_fbopn), norm); col += 1
            ws.write(ri, col, b(db_fbcls), norm); col += 1
            ws.write(ri, col,
                     "OK" if match == "OK" else ("N/F" if "NOT" in action else "DIFF"),
                     ok   if match == "OK" else (warn  if "NOT" in action else fail))
            col += 1
            ws.write(ri, col, action,
                     ok if action == "OK" else (warn if "NOT" in action else fail))
            col += 1

    # Summary sheet
    ws2 = wb.add_worksheet("Summary")
    for c, h in enumerate(["Family", "OK", "MISMATCH", "NOT FOUND", "mc_idx"]):
        ws2.write(0, c, h, hdr)
        ws2.set_column(c, c, [16, 8, 12, 14, 8][c])
    for ri, (fn, mc_idx) in enumerate(families_cfg.items(), start=1):
        fs = fn.replace(" ", "")
        ok_n = mm_n = nf_n = 0
        for rec in comparison_rows:
            a = rec.get(f"{fs}_Action", "")
            if a == "OK":           ok_n += 1
            elif "NOT" in a:        nf_n += 1
            elif "UPDATE" in a:     mm_n += 1
        ws2.write(ri, 0, fn,    norm)
        ws2.write(ri, 1, ok_n,  ok   if ok_n > 0 else norm)
        ws2.write(ri, 2, mm_n,  fail if mm_n > 0 else norm)
        ws2.write(ri, 3, nf_n,  warn if nf_n > 0 else norm)
        ws2.write(ri, 4, mc_idx, norm)

    wb.close()
    print(f"  Written: {out_path}")


# ── Job runner ─────────────────────────────────────────────────────────────────

def run_job(job, udt_em_unit):
    name         = job["name"]
    db_path      = job["db"]
    excel_path   = job["excel"]
    families_cfg = job["families"]        # {family_name: mc_idx}
    out_db       = job.get("out_db",   db_path.replace(".db", "_updated.db"))
    out_xlsx     = job.get("out_xlsx", f"{name}_Comparison.xlsx")

    print("\n" + "=" * 60)
    print(f"  Job: {name}")
    print("=" * 60)

    print(f"\n[1] Reading DB: {db_path}")
    with open(db_path, encoding="utf-8", errors="replace") as f:
        db_text = f.read()
    print(f"    {len(db_text.splitlines())} lines")

    print(f"\n[2] Building instrument map from DB TYPE blocks (unit mapping from UDT)...")
    tag_map = build_tag_map(db_text, udt_em_unit)
    print(f"    {len(tag_map)} instruments discovered")

    print(f"\n[3] Reading Excel: {excel_path}")
    excel_data = parse_excel(excel_path, families_cfg, tag_map)
    for fn, fd in excel_data.items():
        print(f"    {fn}: {len(fd)} instruments")

    print(f"\n[4] Parsing DB type defaults...")
    type_defaults = parse_type_defaults(db_text, tag_map)
    print(f"    {len(type_defaults)} defaults parsed")

    print(f"\n[5] Parsing MachineConfig overrides...")
    mc_overrides_all = {}
    for fn, mc_idx in families_cfg.items():
        ov = parse_mc_overrides(db_text, mc_idx)
        mc_overrides_all[mc_idx] = ov
        print(f"    MC[{mc_idx}] ({fn}): {len(ov)} CW overrides")

    print(f"\n[6] Building comparison...")
    rows = build_comparison(excel_data, type_defaults, mc_overrides_all, tag_map, families_cfg)
    print(f"    {len(rows)} instruments compared")

    print(f"\n    === Mismatch Summary ===")
    for fn in families_cfg:
        fs = fn.replace(" ", "")
        ok_n = mm_n = nf_n = 0
        for rec in rows:
            a = rec.get(f"{fs}_Action", "")
            if a == "OK":       ok_n += 1
            elif "NOT" in a:    nf_n += 1
            elif "UPDATE" in a: mm_n += 1
        print(f"    {fn}: OK={ok_n}  MISMATCH={mm_n}  NOT_FOUND={nf_n}")
        for rec in rows:
            a = rec.get(f"{fs}_Action", "")
            if "UPDATE" in a:
                print(f"      >> {rec['Tag Label']:25s} {a}")

    print(f"\n[7] Writing comparison report: {out_xlsx}")
    write_comparison_xlsx(rows, out_xlsx, families_cfg, name)

    print(f"\n[8] Generating updated DB: {out_db}")
    updated = generate_updated_db(db_text, excel_data, type_defaults,
                                  mc_overrides_all, tag_map, families_cfg)
    with open(out_db, "w", encoding="utf-8") as f:
        f.write(updated)
    print(f"    Written: {out_db}")

    print(f"\n[9] Verifying updated DB...")
    errs = verify_updated_db(updated, excel_data, type_defaults, tag_map, families_cfg)
    if errs == 0:
        print(f"    All ENABLE/EXIST values verified correct.")
    else:
        print(f"    WARNING: {errs} verification error(s)!")

    print(f"\n  COMPLETE: {name}")
    print(f"  Report    : {out_xlsx}")
    print(f"  Updated DB: {out_db}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    try:
        with open(JOBS_FILE, encoding="utf-8") as f:
            cfg = json.load(f)
    except FileNotFoundError:
        print(f"ERROR: {JOBS_FILE} not found.")
        sys.exit(1)

    try:
        with open(UDT_FILE, encoding="utf-8") as f:
            udt_text = f.read()
    except FileNotFoundError:
        print(f"ERROR: {UDT_FILE} not found in working directory.")
        sys.exit(1)
    udt_em_unit = build_em_unit_from_udt(udt_text)
    print(f"[UDT] {len(udt_em_unit)} EM modules mapped from {UDT_FILE}")

    jobs = cfg.get("jobs", [])
    if not jobs:
        print("No jobs defined in jobs.json")
        return

    target = sys.argv[1] if len(sys.argv) > 1 else None

    if target == "--list":
        print("Available jobs:")
        for j in jobs:
            fams = ", ".join(f"{k}=MC[{v}]" for k, v in j.get("families", {}).items())
            enabled = "READY" if j.get("families") else "PENDING (fill families in jobs.json)"
            print(f"  {j['name']:12s}  db={j['db']:30s}  [{enabled}]")
            if fams:
                print(f"  {'':12s}  families: {fams}")
        return

    ran = 0
    for job in jobs:
        if target and job["name"].lower() != target.lower():
            continue
        if not job.get("families"):
            print(f"\nSkipping '{job['name']}': no families configured in jobs.json")
            continue
        run_job(job, udt_em_unit)
        ran += 1

    if ran == 0:
        if target:
            print(f"No job named '{target}' found (or it has no families configured).")
        else:
            print("No jobs were ready to run. Check jobs.json.")

if __name__ == "__main__":
    main()
